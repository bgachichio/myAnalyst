"""NSE end-of-day price adapter.

Source, fixed by Brian: https://www.nse.co.ke/dataservices/market-statistics/

Two honesties are wired into this file rather than left in a comment.

First, **robots.txt is checked before every fetch** and a disallow stops the run.
Second, **the workbook is parsed by column header, not by cell position**, because
a parser pinned to row 7 column D breaks silently the first time a column moves,
and a silent break writes rubbish into the store.

The link-discovery step has NOT been verified against the live page. This
session had no network route to nse.co.ke, so writing selectors against a page
nobody had seen would have been a guess dressed as code. `discover_price_files`
is written generically and is covered by a test that skips until a real saved
copy of the page is dropped into collector/fixtures/. Run it once with a real
sample before this adapter is trusted.
"""
from __future__ import annotations

import datetime as dt
import io
import re
import urllib.parse
import urllib.robotparser
from dataclasses import dataclass

import httpx
import openpyxl

from .htmltable import tables as html_tables
from .store import Quote

SOURCE = "nse.co.ke/dataservices/market-statistics"
PAGE = "https://www.nse.co.ke/dataservices/market-statistics/"
ROBOTS = "https://www.nse.co.ke/robots.txt"
USER_AGENT = "myAnalyst/1.0 (personal research; +https://gachichio.org)"

#: Column headers we accept, lowercased and stripped of punctuation. The NSE has
#: used several wordings over the years; add to these lists rather than editing
#: the parser when a new one appears.
TICKER_HEADERS = ("code", "symbol", "ticker", "counter")
ISIN_HEADERS = ("isin", "isin code", "security code")
SECTOR_HEADERS = ("sector", "segment", "market segment")
NAME_HEADERS = ("name", "company", "security", "company name")
#: The page publishes an explicit download rather than only an inline table.
#: Prefer it: a file is a stable contract, a rendered table is not.
DOWNLOAD_HINT = "daily equity price list"
#: "Statistics as of 02-Sep-2026" sits above the table. Read the date the page
#: states rather than assuming the run date; a run at 00:05 would be a day out.
AS_OF = re.compile(r"statistics\s+as\s+of\s+(\d{1,2})[-\s]([A-Za-z]{3})[-\s](\d{4})", re.I)
MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"], start=1)}
CLOSE_HEADERS = ("day price", "closing price", "close", "price", "last price", "todays price")
VOLUME_HEADERS = ("volume", "shares traded", "traded volume", "no of shares")
#: Words that carry no identity: every counter on the list has them.
NAME_NOISE = frozenset({"LTD", "LIMITED", "PLC", "ORD", "CO", "COMPANY", "GROUP",
                        "HOLDINGS", "KENYA", "THE", "AND"})


class SourceRefused(RuntimeError):
    """robots.txt disallows this path. Not an error to retry or work around."""


class ParseFailed(RuntimeError):
    """The file did not look like a price list. Never write a partial day."""


@dataclass(frozen=True)
class PriceFile:
    url: str
    label: str

    @property
    def is_the_official_download(self) -> bool:
        return DOWNLOAD_HINT in self.label.lower()


def as_of_date(html: str) -> dt.date | None:
    """The trade date the page states for itself, or None if it does not say."""
    match = AS_OF.search(re.sub(r"<[^>]+>", " ", html))
    if not match:
        return None
    day, month, year = match.groups()
    month_number = MONTHS.get(month.lower()[:3])
    return dt.date(int(year), month_number, int(day)) if month_number else None


def _normalise(value: object) -> str:
    """For column headers, where digits are noise."""
    return re.sub(r"[^a-z ]", " ", str(value or "").lower()).strip()


def _label(value: object) -> str:
    """For row labels, where digits are the whole point.

    "NSE 20 Share Index" and "NSE 25 Share Index" differ by one character. Strip
    the digits and they are the same string, and the two indices become
    interchangeable - silently, and for ever, in a store nobody re-reads.
    """
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", str(value or "").lower())).strip()


def robots_allows(url: str, *, client: httpx.Client) -> bool:
    """Ask robots.txt. A fetch that cannot read robots.txt is not attempted."""
    parser = urllib.robotparser.RobotFileParser()
    response = client.get(ROBOTS, timeout=20.0)
    response.raise_for_status()
    parser.parse(response.text.splitlines())
    return parser.can_fetch(USER_AGENT, url)


def make_client() -> httpx.Client:
    return httpx.Client(
        headers={"User-Agent": USER_AGENT, "Accept-Language": "en-GB,en"},
        follow_redirects=True,
        timeout=httpx.Timeout(30.0),
    )


def discover_price_files(html: str, *, base: str = PAGE) -> list[PriceFile]:
    """Find links to price-list workbooks on the market statistics page.

    Generic on purpose: any spreadsheet link whose text or filename mentions a
    price list counts. Verify against a saved copy of the real page before use.
    """
    found: list[PriceFile] = []
    for match in re.finditer(r'<a\b[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', html, re.I | re.S):
        href, label = match.group(1), re.sub(r"<[^>]+>", " ", match.group(2))
        if not re.search(r"\.xlsx?($|\?)", href, re.I):
            continue
        haystack = f"{href} {label}".lower()
        if not any(word in haystack for word in ("price", "pricelist", "equity", "market")):
            continue
        found.append(PriceFile(url=urllib.parse.urljoin(base, href), label=" ".join(label.split())))
    return found


def _header_row(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    """Locate the header row and map the columns we need onto their indices."""
    for index, row in enumerate(rows[:40]):
        cells = [_normalise(c) for c in row]
        found: dict[str, int] = {}
        for key, options in (
            ("ticker", TICKER_HEADERS),
            ("isin", ISIN_HEADERS),
            ("sector", SECTOR_HEADERS),
            ("name", NAME_HEADERS),
            ("close", CLOSE_HEADERS),
            ("volume", VOLUME_HEADERS),
        ):
            for position, cell in enumerate(cells):
                if cell and any(cell == option or cell.startswith(option) for option in options):
                    found.setdefault(key, position)
                    break
        # The published table identifies a security by ISIN and company name,
        # not always by short code, so either identifier will do.
        if ("ticker" in found or "isin" in found or "name" in found) and "close" in found:
            return index, found
    raise ParseFailed(
        "no header row carrying a security identifier and a price column; "
        "the layout changed, or this is not a price list"
    )


def _identify(row: tuple, columns: dict[str, int]) -> tuple[str, str | None]:
    """Return (ticker, isin). Falls back to a slug of the company name."""
    isin = None
    if "isin" in columns:
        candidate = str(row[columns["isin"]] or "").strip().upper()
        if re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}\d", candidate):
            isin = candidate
    if "ticker" in columns:
        ticker = str(row[columns["ticker"]] or "").strip().upper()
        if re.fullmatch(r"[A-Z][A-Z0-9.\-]{1,11}", ticker):
            return ticker, isin
    if "name" in columns:
        # "Williamson Tea Kenya Ltd Ord 5.00" -> WILLIAMSONTEA. Two words, not
        # one: a single word collides across a real board ("Standard Chartered"
        # against "Standard Group"), and a collision silently drops a counter.
        # The ISIN remains the identifier that actually matters.
        all_words = re.findall(r"[A-Za-z]+", str(row[columns["name"]] or ""))
        words = [w for w in all_words if w.upper() not in NAME_NOISE] or all_words
        if words:
            return "".join(words[:2]).upper()[:12], isin
    return "", isin


def _columns(header: list[str]) -> dict[str, int]:
    """Map the columns we need onto their positions, by header name."""
    cells = [_normalise(c) for c in header]
    found: dict[str, int] = {}
    for key, options in (
        ("ticker", TICKER_HEADERS), ("isin", ISIN_HEADERS), ("sector", SECTOR_HEADERS),
        ("name", NAME_HEADERS), ("close", CLOSE_HEADERS), ("volume", VOLUME_HEADERS),
    ):
        for position, cell in enumerate(cells):
            if cell and any(cell == option or cell.startswith(option) for option in options):
                found.setdefault(key, position)
                break
    return found


#: The market summary table is served in the HTML, unlike the equity prices.
#: Rows are matched on their label, tolerantly, because a label is a caption and
#: captions get reworded. Anything unmatched is ignored rather than guessed at.
#: Matched against _label, which keeps digits. Order matters: the most specific
#: alias for each series first.
INDEX_ALIASES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("nse.nasi",    ("nasi", "all share", "allshare")),
    ("nse.nse25",   ("nse 25", "nse25", "25 share")),
    ("nse.banking", ("banking", "bank sector")),
)

#: A quoted index level. Anything outside this band is a turnover figure or a
#: market capitalisation that happens to sit in the same column.
INDEX_BAND = (1.0, 100_000.0)


def parse_market_summary(html: str, obs_date: dt.date) -> list:
    """Read index levels off the market summary table.

    Returns Observations, not Quotes: an index is a series, not a security.
    Imported lazily to keep this module free of a circular import.
    """
    from .store import Observation

    found: dict[str, float] = {}
    for table in html_tables(html):
        if len(table) < 2:
            continue
        header = [_normalise(c) for c in table[0]]
        if not (header and header[0].startswith("name") and any(h.startswith("value") for h in header)):
            continue
        value_at = next(i for i, h in enumerate(header) if h.startswith("value"))

        for row in table[1:]:
            if len(row) <= value_at:
                continue
            label = _label(row[0])
            try:
                value = float(str(row[value_at]).replace(",", "").strip())
            except (TypeError, ValueError):
                continue
            if not INDEX_BAND[0] <= value <= INDEX_BAND[1]:
                continue

            for series_id, aliases in INDEX_ALIASES:
                if not any(a in label for a in aliases):
                    continue
                found.setdefault(series_id, value)
                break

    return [Observation(series_id, obs_date, value, note="market summary")
            for series_id, value in found.items()]


def _from_mapping(row: dict, trade_date: dt.date, sector: str | None) -> Quote | None:
    """Turn one record from a JSON feed into a Quote, matching keys by name."""
    lowered = {_normalise(k): v for k, v in row.items()}

    def pick(options: tuple[str, ...]):
        for key, value in lowered.items():
            if key and any(key == o or key.startswith(o) for o in options):
                return value
        return None

    isin = str(pick(ISIN_HEADERS) or "").strip().upper() or None
    if isin and not re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}\d", isin):
        isin = None

    ticker = str(pick(TICKER_HEADERS) or "").strip().upper()
    if not re.fullmatch(r"[A-Z][A-Z0-9.\-]{1,11}", ticker or ""):
        name = str(pick(NAME_HEADERS) or "")
        all_words = re.findall(r"[A-Za-z]+", name)
        words = [w for w in all_words if w.upper() not in NAME_NOISE] or all_words
        ticker = "".join(words[:2]).upper()[:12] if words else ""
    if not ticker:
        return None

    try:
        close = float(str(pick(CLOSE_HEADERS)).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    if close <= 0:
        return None

    try:
        volume = int(float(str(pick(VOLUME_HEADERS)).replace(",", "").strip()))
    except (TypeError, ValueError):
        volume = None

    row_sector = str(pick(SECTOR_HEADERS) or "").strip().upper() or sector
    return Quote(ticker, trade_date, close, volume, SOURCE, isin=isin, sector=row_sector)


def parse_json_feed(payload: object, trade_date: dt.date, sector: str | None = None) -> list[Quote]:
    """Read quotes from a JSON feed, wherever the records happen to sit.

    A rendered table is markup that can be restyled away; the feed behind it is
    a contract. This accepts a bare list of records, or an object with the list
    under any single key, which covers how these endpoints are usually shaped.
    """
    records: list = []
    if isinstance(payload, list):
        records = payload
    elif isinstance(payload, dict):
        for value in payload.values():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                records = value
                break

    quotes: list[Quote] = []
    seen: set[str] = set()
    for record in records:
        if not isinstance(record, dict):
            continue
        quote = _from_mapping(record, trade_date, sector)
        if quote is None:
            continue
        key = quote.isin or quote.ticker
        if key in seen:
            continue
        quotes.append(quote)
        seen.add(key)
    return quotes


def parse_page_tables(html: str, trade_date: dt.date, sector: str | None = None) -> list[Quote]:
    """Read the price table straight off the page.

    The page renders the day's closes inline, grouped by sector, so this is the
    primary path: it needs no download link to exist and no spreadsheet to
    parse. Every table on the page is considered and the ones that do not carry
    a security identifier and a price are ignored.
    """
    quotes: list[Quote] = []
    seen: set[str] = set()

    for table in html_tables(html):
        if len(table) < 2:
            continue
        columns = _columns(table[0])
        if not (("ticker" in columns or "isin" in columns or "name" in columns) and "close" in columns):
            continue

        for row in table[1:]:
            if len(row) <= max(columns.values()):
                continue
            ticker, isin = _identify(tuple(row), columns)
            key = isin or ticker
            if not ticker or key in seen:
                continue
            try:
                close = float(str(row[columns["close"]]).replace(",", "").strip())
            except (TypeError, ValueError):
                continue          # a suspended counter prints a dash
            if close <= 0:
                continue
            volume = None
            if "volume" in columns:
                try:
                    volume = int(float(str(row[columns["volume"]]).replace(",", "").strip()))
                except (TypeError, ValueError):
                    volume = None
            row_sector = sector
            if "sector" in columns and row[columns["sector"]]:
                row_sector = row[columns["sector"]].strip().upper()
            quotes.append(Quote(ticker, trade_date, close, volume, SOURCE, isin=isin, sector=row_sector))
            seen.add(key)

    return quotes


def parse_workbook(data: bytes, trade_date: dt.date, sector: str | None = None) -> list[Quote]:
    """Parse an NSE price-list workbook into quotes, by column header."""
    book = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    rows = [tuple(r) for r in book[book.sheetnames[0]].iter_rows(values_only=True)]
    start, columns = _header_row(rows)
    book.close()

    quotes: list[Quote] = []
    for row in rows[start + 1:]:
        ticker, isin = _identify(row, columns)
        if not ticker:
            continue
        raw_close = row[columns["close"]]
        try:
            close = float(str(raw_close).replace(",", ""))
        except (TypeError, ValueError):
            continue          # a suspended counter prints a dash, not a price
        if close <= 0:
            continue
        volume = None
        if "volume" in columns:
            try:
                volume = int(float(str(row[columns["volume"]]).replace(",", "")))
            except (TypeError, ValueError):
                volume = None
        row_sector = sector
        if "sector" in columns and row[columns["sector"]]:
            row_sector = str(row[columns["sector"]]).strip().upper()
        quotes.append(Quote(ticker, trade_date, close, volume, SOURCE, isin=isin, sector=row_sector))

    if len(quotes) < 10:
        raise ParseFailed(f"only {len(quotes)} counters parsed; refusing to store a partial day")
    return quotes


#: Below this, the day is partial and must not be stored. The NSE lists well
#: over forty counters; a handful means the scrape caught a fragment.
MINIMUM_COUNTERS = 20


def fetch_latest(
    trade_date: dt.date, *, client: httpx.Client | None = None
) -> tuple[list[Quote], dt.date, str]:
    """Scrape the day's closes. Returns (quotes, trade date, which path worked).

    The page is read first, because the table is rendered inline and needs no
    download link to exist. The linked workbook is the fallback for the day the
    markup changes. Raises rather than returning a partial day: half a price
    list stored is worse than none, because nothing downstream can tell.
    """
    owned = client is None
    client = client or make_client()
    try:
        if not robots_allows(PAGE, client=client):
            raise SourceRefused(f"robots.txt disallows {PAGE}")
        page = client.get(PAGE)
        page.raise_for_status()

        stated = as_of_date(page.text)
        if stated:
            trade_date = stated       # the date the page states for itself

        quotes = parse_page_tables(page.text, trade_date)
        if len(quotes) >= MINIMUM_COUNTERS:
            return quotes, trade_date, "page"

        # Fallback: the page's own "Download Daily Equity Price List".
        files = discover_price_files(page.text)
        if not files:
            raise ParseFailed(
                f"the page yielded {len(quotes)} counters and links no price-list workbook; "
                "the markup has changed"
            )
        newest = next((f for f in files if f.is_the_official_download), files[0])
        if not robots_allows(newest.url, client=client):
            raise SourceRefused(f"robots.txt disallows {newest.url}")
        workbook = client.get(newest.url)
        workbook.raise_for_status()
        from_file = parse_workbook(workbook.content, trade_date)
        if len(from_file) < MINIMUM_COUNTERS:
            raise ParseFailed(f"only {len(from_file)} counters from the workbook; refusing a partial day")
        return from_file, trade_date, "workbook"
    finally:
        if owned:
            client.close()
